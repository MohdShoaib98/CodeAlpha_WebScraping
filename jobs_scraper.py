import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Categories to scrape
categories = [
    "https://remoteok.com/remote-dev-jobs",
    "https://remoteok.com/remote-design-jobs",
    "https://remoteok.com/remote-marketing-jobs",
    "https://remoteok.com/remote-writing-jobs",
    "https://remoteok.com/remote-sales-jobs"
]

def scrape_remoteok(url):
    """
    Scrapes job data from a RemoteOK category page.

    Args:
        url (str): The URL of the RemoteOK category page.

    Returns:
        list: A list of dictionaries, where each dictionary represents a job.
    """
    try:
        response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"})
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")

        jobs = []
        table = soup.find("table", id="jobsboard")
        if table:
            rows = table.find_all("tr", class_="job")
            for job in rows:
                # Title extraction
                title = job.get("data-position") or (job.find("h2", itemprop="title").text.strip() if job.find("h2", itemprop="title") else "Title Not Found")

                # Company extraction
                company = job.get("data-company") or (job.find("h3", itemprop="name").text.strip() if job.find("h3", itemprop="name") else "Company Not Found")

                # Location extraction
                location = "Worldwide"
                company_td = job.find("td", class_="company")
                if company_td:
                    location_div = company_td.find("div", class_="location")
                    location = location_div.text.strip() if location_div and location_div.text.strip() else "Worldwide" # Simplified

                # Link extraction
                link = f"https://remoteok.com{job.get('data-href', '')}" # Simplified

                # Salary extraction
                salary_element = job.find("div", class_="salary")
                salary = salary_element.text.strip() if salary_element else "Not specified"

                # Category extraction
                category = url.split("/")[-1].replace("remote-", "").replace("-jobs", "").capitalize()

                jobs.append({
                    "Category": category,
                    "Title": title,
                    "Company": company,
                    "Location": location,
                    "Salary": salary,
                    "Link": link
                })
        return jobs
    except requests.exceptions.RequestException as e:
        print(f"Error fetching {url}: {e}")
        return []
    except Exception as e:
        print(f"Error parsing {url}: {e}")
        return []

def create_excel(jobs, file_name="jobs_multi.xlsx"):
    """
    Creates an Excel file from the scraped job data and applies styling.

    Args:
        jobs (list): A list of job dictionaries.
        file_name (str): The name of the Excel file to create.
    """
    if not jobs:
        print("No jobs to save.")
        return

    df = pd.DataFrame(jobs)
    df.to_excel(file_name, index=False)

    wb = load_workbook(file_name)
    ws = wb.active

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for col_num, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_num)].width = max(len(col_name) + 5, 20)

    for row in range(2, ws.max_row + 1):
        fill_color = "D9E1F2" if row % 2 == 0 else "FFFFFF"
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.border = thin_border

    ws.auto_filter.ref = ws.dimensions
    wb.save(file_name)

    print(f"✅ Saved {len(jobs)} jobs to {file_name} with styling")

# Main execution
if __name__ == "__main__":
    all_jobs = []
    print("Fetching jobs from RemoteOK categories...")

    for url in categories:
        jobs = scrape_remoteok(url)
        all_jobs.extend(jobs)

    if all_jobs:
        create_excel(all_jobs)
    else:
        print("No jobs were scraped.")
